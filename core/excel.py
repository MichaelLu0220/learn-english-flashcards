import openpyxl
from pathlib import Path
from typing import List, Optional, Tuple
from core.models import WordEntry, ExcelColumn, ExcelSchema

ROOT_DIR = Path(__file__).parent.parent / "data"

class ExcelReader:
    """智慧 Excel 讀取器"""
    
    # 欄位關鍵字對應
    FIELD_KEYWORDS = {
        "word": ["word", "單字", "英文", "english", "vocabulary", "vocab", "字彙"],
        "meaning": ["meaning", "意思", "中文", "解釋", "definition", "chinese", "翻譯", "translation"],
        "example": ["example", "例句", "sentence", "造句", "用法", "usage"],
    }
    
    def __init__(self, file_path: str, sheet_name: str = ""):
        self.file_path = ROOT_DIR / file_path
        self.sheet_name = sheet_name
        self.workbook = None
        self.worksheet = None
        self.schema: Optional[ExcelSchema] = None
        self.word_entries: List[WordEntry] = []
        
    def load(self) -> Tuple[bool, str]:
        """
        載入 Excel 檔案
        Returns: (成功與否, 訊息)
        """
        try:
            if not self.file_path.exists():
                return False, f"找不到檔案: {self.file_path}"
            
            self.workbook = openpyxl.load_workbook(self.file_path, read_only=True)
            
            # 選擇工作表
            if self.sheet_name and self.sheet_name in self.workbook.sheetnames:
                self.worksheet = self.workbook[self.sheet_name]
            else:
                self.worksheet = self.workbook.active
                self.sheet_name = self.worksheet.title
            
            # 分析結構並載入資料
            self.schema = self._analyze_schema()
            self._parse_entries()
            
            return True, f"載入成功：{len(self.word_entries)} 個單字"
            
        except Exception as e:
            return False, f"載入失敗: {e}"
    
    def _analyze_schema(self) -> ExcelSchema:
        """分析 Excel 結構"""
        # 讀取前兩列來判斷
        rows = list(self.worksheet.iter_rows(min_row=1, max_row=2, values_only=True))
        
        if not rows:
            return ExcelSchema(has_header=False, columns=[])
        
        first_row = rows[0]
        schema = ExcelSchema(has_header=False, columns=[])
        
        # 嘗試辨識第一列是否為標題
        header_matches = self._try_match_header(first_row)
        
        if header_matches:
            # 有標題列
            schema.has_header = True
            schema.columns = header_matches
            print(f"偵測到標題列: {[c.name for c in schema.columns]}")
        else:
            # 無標題，按順序假設欄位
            schema.has_header = False
            schema.columns = self._create_default_columns(len(first_row))
            print(f"無標題列，使用預設順序（共 {len(first_row)} 欄）")
        
        return schema
    
    def _try_match_header(self, row: tuple) -> Optional[List[ExcelColumn]]:
        """嘗試匹配標題列"""
        columns = []
        matched_fields = set()
        
        for idx, cell in enumerate(row):
            if cell is None:
                continue
                
            cell_str = str(cell).strip().lower()
            matched_field = None
            
            # 嘗試匹配每個欄位的關鍵字
            for field_name, keywords in self.FIELD_KEYWORDS.items():
                if field_name in matched_fields:
                    continue
                    
                for keyword in keywords:
                    if keyword in cell_str:
                        matched_field = field_name
                        matched_fields.add(field_name)
                        break
                
                if matched_field:
                    break
            
            if matched_field:
                columns.append(ExcelColumn(
                    index=idx,
                    name=str(cell).strip(),
                    field=matched_field
                ))
        
        # 至少要匹配到 word 或 meaning 才算有效標題
        if "word" in matched_fields or "meaning" in matched_fields:
            return columns
        
        return None
    
    def _create_default_columns(self, col_count: int) -> List[ExcelColumn]:
        """建立預設欄位順序"""
        default_order = ["word", "meaning", "example"]
        columns = []
        
        for idx in range(min(col_count, len(default_order))):
            columns.append(ExcelColumn(
                index=idx,
                name=default_order[idx],
                field=default_order[idx]
            ))
        
        return columns
    
    def _parse_entries(self):
        """解析單字條目"""
        self.word_entries.clear()
        
        if not self.schema or not self.schema.columns:
            return
        
        # 決定起始列
        start_row = 2 if self.schema.has_header else 1
        
        # 取得欄位索引
        word_idx = self.schema.get_field_index("word")
        meaning_idx = self.schema.get_field_index("meaning")
        example_idx = self.schema.get_field_index("example")
        
        for row in self.worksheet.iter_rows(min_row=start_row, values_only=True):
            # 跳過空列
            if not any(cell is not None and str(cell).strip() for cell in row):
                continue
            
            entry = WordEntry(
                word=self._get_cell(row, word_idx),
                meaning=self._get_cell(row, meaning_idx),
                example=self._get_cell(row, example_idx),
            )
            
            if entry.has_content():
                self.word_entries.append(entry)
    
    def _get_cell(self, row: tuple, idx: Optional[int]) -> str:
        """安全取得儲存格內容"""
        if idx is None or idx >= len(row) or row[idx] is None:
            return ""
        return str(row[idx]).strip()
    
    def reload(self) -> Tuple[bool, str]:
        """重新載入"""
        return self.load()
    
    def get_sheet_names(self) -> List[str]:
        """取得所有工作表名稱"""
        if self.workbook:
            return self.workbook.sheetnames
        return []